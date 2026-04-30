# -*- coding: utf-8 -*-
"""
TaxOpt v0.2 검증팀 손계산 양식 생성 스크립트 (v2)
v0.1 04_test_cases_manual.xlsx의 형식·스타일을 그대로 따른다.

시트 구조:
  0_규칙          — 작성 규칙·표준 변수명·색상 범례
  1_TC-006        — 1세대1주택 비과세 + 12억 이하
  2_TC-007        — 1세대1주택 + 12억 초과 (안분 + 표 2 64%)
  3_TC-008        — 다주택 일반과세 + 표 1 (보유 12년)
  4_TC-009        — 1세대1주택 + 표 2 최대 80% (안분 14억)
  5_TC-010        — 일시적 2주택 (적용 안 함)
  6_세율표         — 8단계 누진세율표
  7_장특공표       — 표 1·표 2

색상 범례 (v0.1 양식 그대로):
  파란 글자 (FF0000FF): 입력값
  검정 글자 (FF000000): 수식·계산 결과
  연두 배경 (FFE2EFDA): 최종 결과 셀
  진한 파랑 배경 (FF1F4E78) + 흰 글자: 시트 제목
  중간 파랑 배경 (FF2E75B6) + 흰 글자: 섹션 헤더
  연한 파랑 배경 (FFDEEBF7): 라벨 셀
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 스타일 (v0.1 양식 그대로)
# ============================================================
TITLE_FILL    = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FILL   = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
LABEL_FILL    = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
RESULT_FILL   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

WHITE_BOLD    = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
BLACK_BOLD    = Font(name="맑은 고딕", size=11, bold=True, color="000000")
BLACK_NORMAL  = Font(name="맑은 고딕", size=11, bold=False, color="000000")
BLUE_INPUT    = Font(name="맑은 고딕", size=11, bold=False, color="0000FF")  # 입력값 (작성자가 채움)
TITLE_FONT    = Font(name="맑은 고딕", size=13, bold=True, color="FFFFFF")

THIN          = Side(style='thin', color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT          = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT         = Alignment(horizontal='right',  vertical='center', wrap_text=False)

# ============================================================
# 헬퍼
# ============================================================
def set_title(ws, coord, text, span=None):
    ws[coord] = text
    ws[coord].font = TITLE_FONT
    ws[coord].fill = TITLE_FILL
    ws[coord].alignment = CENTER
    if span:
        ws.merge_cells(span)

def set_section(ws, coord, text, span=None):
    ws[coord] = text
    ws[coord].font = WHITE_BOLD
    ws[coord].fill = HEADER_FILL
    ws[coord].alignment = LEFT
    if span:
        ws.merge_cells(span)

def set_label(ws, coord, text):
    ws[coord] = text
    ws[coord].font = BLACK_BOLD
    ws[coord].fill = LABEL_FILL
    ws[coord].alignment = LEFT
    ws[coord].border = BORDER

def set_value(ws, coord, value, kind="text"):
    """kind: text, input(파랑), formula(검정), result(연두 배경)"""
    ws[coord] = value
    ws[coord].border = BORDER
    if kind == "input":
        ws[coord].font = BLUE_INPUT
        ws[coord].alignment = RIGHT if isinstance(value, (int, float)) else LEFT
        if isinstance(value, int) and abs(value) >= 1000:
            ws[coord].number_format = '#,##0'
    elif kind == "formula":
        ws[coord].font = BLACK_NORMAL
        ws[coord].alignment = RIGHT
        ws[coord].number_format = '#,##0'
    elif kind == "result":
        ws[coord].font = BLACK_BOLD
        ws[coord].fill = RESULT_FILL
        ws[coord].alignment = RIGHT
        ws[coord].number_format = '#,##0'
    else:
        ws[coord].font = BLACK_NORMAL
        ws[coord].alignment = LEFT

# ============================================================
# 시트 0: 규칙
# ============================================================
def make_rules_sheet(wb):
    ws = wb.create_sheet("0_규칙")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12

    set_title(ws, 'A1', "TaxOpt v0.2 수기 정답 작성 규칙", "A1:C1")

    # 1. 작성 규칙
    set_section(ws, 'A3', "1. 작성 규칙", "A3:C3")
    rules = [
        ("(1) 변수명",      "한글 라벨 옆에 영문 변수명 병기. 영문은 코드 변수명과 일치시킨다."),
        ("(2) 단위",        "모든 금액은 원 단위 정수. 1억 = 100,000,000 (콤마 표기는 가독성용 허용)."),
        ("(3) 날짜",        "ISO 형식 YYYY-MM-DD. 예: 2026-08-31"),
        ("(4) 개월·연",     "정수. 보유기간 3년 5개월 → 41개월"),
        ("(5) 법령 인용",   "「소득세법 제○○조 제○항 제○호」 형식. 시행령은 「소득세법 시행령 제○○조」"),
        ("(6) 세율 적용",   "8단계 세율표 중 어느 구간인지 명시. 적용 산식까지 풀어 쓴다."),
        ("(7) 허용 오차",   "1원 단위 정확 일치. 5원 이내 차이는 반올림 차이로 간주(허용)."),
    ]
    for i, (k, v) in enumerate(rules):
        r = 4 + i
        set_label(ws, f'A{r}', k)
        ws[f'B{r}'] = v
        ws[f'B{r}'].font = BLACK_NORMAL
        ws[f'B{r}'].alignment = LEFT
        ws[f'B{r}'].border = BORDER
        ws.merge_cells(f'B{r}:C{r}')
        ws.row_dimensions[r].height = 30

    # 2. 표준 변수명 표
    set_section(ws, 'A12', "2. 표준 변수명 표", "A12:C12")
    set_label(ws, 'A13', "한글")
    set_label(ws, 'B13', "영문 변수명 (코드와 일치)")
    set_label(ws, 'C13', "단위")

    vars_map = [
        ("양도가액",                                 "salePrice",             "원"),
        ("취득가액",                                 "acquisitionPrice",      "원"),
        ("필요경비",                                 "necessaryExpense",      "원"),
        ("양도차익",                                 "transferGain",          "원"),
        ("과세대상 양도차익 (고가주택 안분 후)",     "taxableGain",           "원"),
        ("장기보유특별공제액",                       "longTermDeduction",     "원"),
        ("양도소득금액",                             "capitalGainIncome",     "원"),
        ("양도소득 기본공제",                        "basicDeduction",        "원"),
        ("과세표준",                                 "taxBase",               "원"),
        ("산출세액",                                 "calculatedTax",         "원"),
        ("지방소득세",                               "localIncomeTax",        "원"),
        ("총 납부세액",                              "totalTax",              "원"),
        ("세후 매각금액",                            "netAfterTaxSaleAmount", "원"),
        ("보유기간",                                 "holdingPeriodMonths",   "개월"),
        ("거주기간",                                 "residencePeriodMonths", "개월"),
        ("1세대1주택 비과세 적용 여부",              "is1Se1House",           "—"),
        ("고가주택 여부 (12억 초과)",                "isHighValueHouse",      "—"),
        ("적용 장특공 표 (1 또는 2)",                "appliedDeductionTable", "—"),
        ("일시적 2주택 여부 (v0.2 미적용)",          "isOneTimeTwoHouses",    "—"),
        ("1세대 보유 주택 수",                       "householdHouseCount",   "개"),
    ]
    for i, (k, e, u) in enumerate(vars_map):
        r = 14 + i
        set_value(ws, f'A{r}', k)
        set_value(ws, f'B{r}', e)
        set_value(ws, f'C{r}', u)

    # 3. 색상 범례
    base_r = 14 + len(vars_map) + 1
    set_section(ws, f'A{base_r}', "3. 색상 범례", f"A{base_r}:C{base_r}")
    legend = [
        ("파란 글자",     "작성자가 입력하는 값 (입력값·하드코딩 숫자)",   BLUE_INPUT,   None),
        ("검정 글자",     "수식·계산 결과",                                BLACK_NORMAL, None),
        ("연두 배경",     "최종 결과값 셀",                                BLACK_BOLD,   RESULT_FILL),
        ("진한 파랑 배경","시트 제목·섹션 제목",                           WHITE_BOLD,   TITLE_FILL),
        ("연한 파랑 배경","라벨·항목명",                                   BLACK_BOLD,   LABEL_FILL),
    ]
    for i, (k, v, font, fill) in enumerate(legend):
        r = base_r + 1 + i
        ws[f'A{r}'] = k
        ws[f'A{r}'].font = font
        if fill:
            ws[f'A{r}'].fill = fill
        ws[f'A{r}'].alignment = CENTER
        ws[f'A{r}'].border = BORDER
        ws[f'B{r}'] = v
        ws[f'B{r}'].font = BLACK_NORMAL
        ws[f'B{r}'].alignment = LEFT
        ws[f'B{r}'].border = BORDER
        ws.merge_cells(f'B{r}:C{r}')

    # 4. v0.2 신규 안내
    base_r2 = base_r + 1 + len(legend) + 1
    set_section(ws, f'A{base_r2}', "4. v0.2에서 신규 적용되는 항목", f"A{base_r2}:C{base_r2}")
    new_items = [
        ("(1) 1세대1주택 비과세", "양도가액 12억 이하 → 전액 비과세 (totalTax = 0). 12억 초과 → 안분 과세."),
        ("(2) 고가주택 안분",      "과세대상 양도차익 = 양도차익 × (양도가액 - 12억) / 양도가액"),
        ("(3) 장특공 표 2",        "1세대1주택: 보유 ×4%/년 + 거주 ×4%/년, 최대 80%"),
        ("(4) 일시적 2주택",       "v0.2 미적용. issueFlag(ONE_TIME_2HOUSES_NOT_APPLIED)만 발동, 산식 분기 없음."),
    ]
    for i, (k, v) in enumerate(new_items):
        r = base_r2 + 1 + i
        set_label(ws, f'A{r}', k)
        ws[f'B{r}'] = v
        ws[f'B{r}'].font = BLACK_NORMAL
        ws[f'B{r}'].alignment = LEFT
        ws[f'B{r}'].border = BORDER
        ws.merge_cells(f'B{r}:C{r}')
        ws.row_dimensions[r].height = 30

# ============================================================
# 공통 — TC 시트 작성 (v0.1 1_TC-001 형식 그대로)
# ============================================================
def make_tc_sheet(wb, sheet_name, tc_id, intent, version, inputs, expected_steps,
                   law_refs, assumptions, issue_flags_text=None):
    """
    inputs: list of (label_kr, var_eng, value, unit, note)  — 입력값
    expected_steps: list of (var_eng, formula_text, claude_estimated_value, note)  — 단계별 산식
    law_refs: list of (law_text, description)
    assumptions: list of strings
    issue_flags_text: list of strings (each issueFlag 발동 안내)
    """
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 50

    # 헤더
    set_title(ws, 'A1', f"케이스 {tc_id} — {intent}", "A1:D1")

    set_label(ws, 'A2', "케이스 ID");      set_value(ws, 'B2', tc_id)
    set_label(ws, 'A3', "버전 적용");      set_value(ws, 'B3', version)
    set_label(ws, 'A4', "작성자");         set_value(ws, 'B4', "")
    set_label(ws, 'A5', "작성일");         set_value(ws, 'B5', "")
    set_label(ws, 'A6', "검토자");         set_value(ws, 'B6', "")
    set_label(ws, 'A7', "상태");           set_value(ws, 'B7', "작성 중")

    # 1. 입력값
    set_section(ws, 'A9', "1. 입력값 (Given)", "A9:D9")
    set_label(ws, 'A10', "항목"); set_label(ws, 'B10', "값"); set_label(ws, 'C10', "단위"); set_label(ws, 'D10', "비고")

    r = 11
    for label_kr, var_eng, value, unit, note in inputs:
        if var_eng:
            set_value(ws, f'A{r}', f"{label_kr} ({var_eng})")
        else:
            set_value(ws, f'A{r}', label_kr)
        set_value(ws, f'B{r}', value, kind="input")
        set_value(ws, f'C{r}', unit)
        set_value(ws, f'D{r}', note or "")
        r += 1

    # 2. 적용 법령
    r += 1
    set_section(ws, f'A{r}', "2. 적용 법령 (Reference)", f"A{r}:D{r}")
    r += 1
    for law, desc in law_refs:
        set_value(ws, f'A{r}', law)
        ws[f'B{r}'] = desc
        ws[f'B{r}'].font = BLACK_NORMAL
        ws[f'B{r}'].alignment = LEFT
        ws[f'B{r}'].border = BORDER
        ws.merge_cells(f'B{r}:D{r}')
        r += 1

    # 3. 계산 단계
    r += 1
    set_section(ws, f'A{r}', "3. 계산 단계 (Step-by-step) — 검증팀이 산식·결과값을 작성", f"A{r}:D{r}")
    r += 1
    set_label(ws, f'A{r}', "변수명"); set_label(ws, f'B{r}', "산식 (검증팀 작성)")
    set_label(ws, f'C{r}', "결과값"); set_label(ws, f'D{r}', "비고 / Claude 명세서 산출 (참고)")
    r += 1
    for var_eng, formula_text, claude_value, note in expected_steps:
        set_value(ws, f'A{r}', var_eng)
        # 산식: 작성자가 입력 (파란 글자) — 빈칸으로 두고 작성자 작성
        if formula_text:
            ws[f'B{r}'] = formula_text
            ws[f'B{r}'].font = BLUE_INPUT
            ws[f'B{r}'].alignment = LEFT
            ws[f'B{r}'].border = BORDER
        else:
            set_value(ws, f'B{r}', "")
        # 결과값: 작성자 (파란 글자), Claude 추정값은 비고에
        set_value(ws, f'C{r}', "", kind="input")
        # 비고에 Claude 추정값 안내
        note_full = note or ""
        if claude_value is not None and claude_value != "":
            if note_full:
                note_full = f"{note_full} / Claude 추정값: {claude_value:,}원" if isinstance(claude_value, int) else f"{note_full} / Claude 추정값: {claude_value}"
            else:
                note_full = f"Claude 추정값: {claude_value:,}원" if isinstance(claude_value, int) else f"Claude 추정값: {claude_value}"
        set_value(ws, f'D{r}', note_full)
        r += 1

    # 4. 결과값 요약 (연두 배경)
    r += 1
    set_section(ws, f'A{r}', "4. 결과값 (요약)", f"A{r}:D{r}")
    r += 1
    set_label(ws, f'A{r}', "totalTax (최종 납부세액)")
    ws[f'B{r}'] = ""; ws[f'B{r}'].fill = RESULT_FILL; ws[f'B{r}'].border = BORDER
    ws[f'B{r}'].font = BLACK_BOLD; ws[f'B{r}'].alignment = RIGHT; ws[f'B{r}'].number_format = '#,##0'
    ws.merge_cells(f'B{r}:D{r}')
    r += 1
    set_label(ws, f'A{r}', "netAfterTaxSaleAmount (세후 매각금액)")
    ws[f'B{r}'] = ""; ws[f'B{r}'].fill = RESULT_FILL; ws[f'B{r}'].border = BORDER
    ws[f'B{r}'].font = BLACK_BOLD; ws[f'B{r}'].alignment = RIGHT; ws[f'B{r}'].number_format = '#,##0'
    ws.merge_cells(f'B{r}:D{r}')

    # 5. 가정·주의사항
    r += 2
    set_section(ws, f'A{r}', "5. 가정 · 주의사항 (Assumptions / Caveats)", f"A{r}:D{r}")
    r += 1
    for a in assumptions:
        ws[f'A{r}'] = a
        ws[f'A{r}'].font = BLACK_NORMAL
        ws[f'A{r}'].alignment = LEFT
        ws[f'A{r}'].border = BORDER
        ws.merge_cells(f'A{r}:D{r}')
        r += 1

    # 6. issueFlag (작성자가 발동 의도 검토)
    r += 1
    set_section(ws, f'A{r}', "6. issueFlag (발동 검토)", f"A{r}:D{r}")
    r += 1
    if issue_flags_text:
        for ift in issue_flags_text:
            ws[f'A{r}'] = ift
            ws[f'A{r}'].font = BLACK_NORMAL
            ws[f'A{r}'].alignment = LEFT
            ws[f'A{r}'].border = BORDER
            ws.merge_cells(f'A{r}:D{r}')
            r += 1
    else:
        ws[f'A{r}'] = "(법령 해석상 불명확하거나 추가 검토가 필요한 항목 — 없으면 '없음')"
        ws[f'A{r}'].font = BLACK_NORMAL
        ws[f'A{r}'].alignment = LEFT
        ws[f'A{r}'].border = BORDER
        ws.merge_cells(f'A{r}:D{r}')

# ============================================================
# 시트 6·7: 세율표·장특공표 (v0.1과 동일)
# ============================================================
def make_rate_sheet(wb):
    ws = wb.create_sheet("6_세율표")
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 38

    set_title(ws, 'A1', "소득세법 제55조 제1항 기본세율표 (양도소득세에도 적용)", "A1:E1")

    headers = ["과세표준 하한 (원)", "과세표준 상한 (원)", "세율", "산식 (법령 본문)", "누진공제 환산 (참고)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = BLACK_BOLD
        cell.fill = LABEL_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    data = [
        (0,           14_000_000,  0.06, "과세표준의 6%",                                   "(누진공제 없음)"),
        (14_000_000,  50_000_000,  0.15, "84만원 + 1,400만원 초과액의 15%",                "과세표준 × 15% - 1,260,000원"),
        (50_000_000,  88_000_000,  0.24, "624만원 + 5,000만원 초과액의 24%",               "과세표준 × 24% - 5,760,000원"),
        (88_000_000,  150_000_000, 0.35, "1,536만원 + 8,800만원 초과액의 35%",             "과세표준 × 35% - 15,440,000원"),
        (150_000_000, 300_000_000, 0.38, "3,706만원 + 1억5천만원 초과액의 38%",            "과세표준 × 38% - 19,940,000원"),
        (300_000_000, 500_000_000, 0.40, "9,406만원 + 3억원 초과액의 40%",                 "과세표준 × 40% - 25,940,000원"),
        (500_000_000, 1_000_000_000, 0.42, "1억7,406만원 + 5억원 초과액의 42%",            "과세표준 × 42% - 35,940,000원"),
        (1_000_000_000, "초과",     0.45, "3억8,406만원 + 10억원 초과액의 45%",             "과세표준 × 45% - 65,940,000원"),
    ]
    for i, row in enumerate(data):
        r = 4 + i
        for col_idx, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.font = BLACK_NORMAL
            cell.border = BORDER
            if col_idx in (1, 2):
                cell.alignment = RIGHT
                if isinstance(v, int):
                    cell.number_format = '#,##0'
            elif col_idx == 3:
                cell.alignment = CENTER
                cell.number_format = '0.00'
            else:
                cell.alignment = LEFT

    # 사용 예시
    set_section(ws, 'A14', "사용 예시 — 과세표준 2억원의 경우", "A14:E14")
    ws['A15'] = "방식 1 (법령 본문)"; ws['A15'].font = BLACK_BOLD; ws['A15'].fill = LABEL_FILL; ws['A15'].border = BORDER
    ws['B15'] = "3,706만원 + (200,000,000 - 150,000,000) × 38% = 56,060,000원"
    ws['B15'].font = BLACK_NORMAL; ws['B15'].alignment = LEFT; ws['B15'].border = BORDER
    ws.merge_cells('B15:E15')

    ws['A16'] = "방식 2 (누진공제)"; ws['A16'].font = BLACK_BOLD; ws['A16'].fill = LABEL_FILL; ws['A16'].border = BORDER
    ws['B16'] = "200,000,000 × 38% - 19,940,000 = 56,060,000원"
    ws['B16'].font = BLACK_NORMAL; ws['B16'].alignment = LEFT; ws['B16'].border = BORDER
    ws.merge_cells('B16:E16')


def make_deduction_sheet(wb):
    ws = wb.create_sheet("7_장특공표")
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 12

    set_title(ws, 'A1', "소득세법 제95조 제2항 장기보유특별공제 표", "A1:D1")

    set_section(ws, 'A3', "표1: 일반 (다주택 중과 미적용)",          "A3:B3")
    set_section(ws, 'C3', "표2: 1세대1주택 (보유 + 거주 합산)",      "C3:D3")

    set_label(ws, 'A4', "보유기간"); set_label(ws, 'B4', "공제율")
    set_label(ws, 'C4', "기간 (보유 또는 거주)"); set_label(ws, 'D4', "공제율 (각각)")

    table1 = [
        ("3년 이상 4년 미만", 0.06),
        ("4년 이상 5년 미만", 0.08),
        ("5년 이상 6년 미만", 0.10),
        ("6년 이상 7년 미만", 0.12),
        ("7년 이상 8년 미만", 0.14),
        ("8년 이상 9년 미만", 0.16),
        ("9년 이상 10년 미만",0.18),
        ("10년 이상 11년 미만",0.20),
        ("11년 이상 12년 미만",0.22),
        ("12년 이상 13년 미만",0.24),
        ("13년 이상 14년 미만",0.26),
        ("14년 이상 15년 미만",0.28),
        ("15년 이상",         0.30),
    ]
    table2 = [
        ("3년 이상 4년 미만", 0.12),
        ("4년 이상 5년 미만", 0.16),
        ("5년 이상 6년 미만", 0.20),
        ("6년 이상 7년 미만", 0.24),
        ("7년 이상 8년 미만", 0.28),
        ("8년 이상 9년 미만", 0.32),
        ("9년 이상 10년 미만",0.36),
        ("10년 이상",         0.40),
    ]
    for i, (k, v) in enumerate(table1):
        r = 5 + i
        set_value(ws, f'A{r}', k)
        cell = ws[f'B{r}']
        cell.value = v
        cell.font = BLACK_NORMAL
        cell.alignment = CENTER
        cell.border = BORDER
        cell.number_format = '0.00'
    for i, (k, v) in enumerate(table2):
        r = 5 + i
        set_value(ws, f'C{r}', k)
        cell = ws[f'D{r}']
        cell.value = v
        cell.font = BLACK_NORMAL
        cell.alignment = CENTER
        cell.border = BORDER
        cell.number_format = '0.00'

    # 주의사항
    base = 5 + len(table1) + 1
    set_section(ws, f'A{base}', "주의사항", f"A{base}:D{base}")
    notes = [
        "1. 표1은 일반 양도. 다주택 중과 적용 대상 주택은 장특공 적용 배제 (소득세법 제95조 ②).",
        "2. 표2는 1세대1주택 보유분과 거주분을 각각 적용하여 합산 (최대 80%).",
        "3. v0.2는 표1·표2 모두 적용. 1세대1주택(householdHouseCount=1) → 표2, 그 외 → 표1.",
        "4. 보유기간은 취득일~양도일까지의 기간을 만(滿) 단위로 계산.",
        "5. 거주기간은 거주일수 합계를 만(滿) 개월 단위로 계산 (소득세법 시행령 제159조의4).",
    ]
    for i, note in enumerate(notes):
        r = base + 1 + i
        ws[f'A{r}'] = note
        ws[f'A{r}'].font = BLACK_NORMAL
        ws[f'A{r}'].alignment = LEFT
        ws[f'A{r}'].border = BORDER
        ws.merge_cells(f'A{r}:D{r}')

# ============================================================
# TC별 데이터
# ============================================================

# 공통 법령 인용
COMMON_LAW_REFS = [
    ("소득세법 제55조 제1항",        "기본세율표 (양도소득세에도 적용)"),
    ("소득세법 제89조 제1항 제3호",  "1세대1주택 비과세 (12억 이하)"),
    ("소득세법 제95조 제2항",        "장기보유특별공제 — 표1·표2"),
    ("소득세법 제95조 제3항",        "고가주택 안분 — 12억 초과분 과세"),
    ("소득세법 시행령 제154조",      "1세대1주택 비과세 요건"),
    ("소득세법 시행령 제159조의3",   "장특공 표1 (일반)"),
    ("소득세법 시행령 제159조의4",   "장특공 표2 (1세대1주택, 보유+거주)"),
    ("소득세법 시행령 제160조",      "고가주택 안분 산식"),
    ("소득세법 제103조",             "양도소득 기본공제 250만원"),
    ("소득세법 제104조 제1항",       "양도소득세율"),
    ("지방세법 제103조의3",          "지방소득세율 10%"),
]

TC_DATA = {
    "TC-006": {
        "intent": "1세대1주택 비과세 + 12억 이하 (전액 비과세)",
        "version": "v0.2",
        "inputs": [
            ("기준연도",           "baseYear",              2026,            "—",  ""),
            ("세대 구성원 수",     "householdMembers",      2,               "명", ""),
            ("기본공제 사용 여부", "basicDeductionUsed",    "false",         "—",  ""),
            ("1세대 보유 주택 수", "householdHouseCount",   1,               "개", "1세대1주택 분기 진입"),
            ("일시적 2주택 여부",  "isOneTimeTwoHouses",    "false",         "—",  ""),
            ("주택 ID",            None,                    "A",             "—",  ""),
            ("취득일",             "acquisitionDate",       "2021-04-30",    "—",  ""),
            ("취득가액",           "acquisitionPrice",      600_000_000,     "원", ""),
            ("필요경비",           "necessaryExpense",      15_000_000,      "원", ""),
            ("취득시 조정대상지역", "acquisitionRegulated", "false",         "—",  "비조정대상 → 거주요건 면제"),
            ("거주기간",           "residenceMonths",       60,              "개월", "5년"),
            ("현재 거주 여부",     "livingNow",             "true",          "—",  ""),
            ("양도예정일",         "expectedSaleDate",      "2026-08-31",    "—",  "중과 유예 종료 후"),
            ("양도가액",           "expectedSalePrice",     1_000_000_000,   "원", "10억 (12억 이하)"),
            ("양도시 조정대상지역", "saleRegulated",        "false",         "—",  ""),
        ],
        "expected_steps": [
            ("transferGain",          "salePrice - acquisitionPrice - necessaryExpense", 385_000_000, "양도차익"),
            ("is1Se1House",           "householdHouseCount=1 AND 거주요건 충족",          "true",      "1세대1주택 분기 진입"),
            ("isHighValueHouse",      "salePrice ≥ 1,200,000,000 ?",                    "false",     "12억 이하 → 전액 비과세"),
            ("terminateAt2",          "비과세 + 12억 이하 → 단계 2 종료",                  "true",      "단계 3~12 스킵"),
            ("taxableGain",           "단계 2 종료 시 0",                                 0,           ""),
            ("longTermDeduction",     "(스킵)",                                          0,           ""),
            ("capitalGainIncome",     "(스킵)",                                          0,           ""),
            ("basicDeduction",        "(스킵)",                                          0,           ""),
            ("taxBase",               "(스킵)",                                          0,           ""),
            ("calculatedTax",         "(스킵)",                                          0,           ""),
            ("localIncomeTax",        "(스킵)",                                          0,           ""),
            ("totalTax",              "0 (전액 비과세)",                                  0,           "최종"),
            ("netAfterTaxSaleAmount", "salePrice - totalTax",                            1_000_000_000, "최종"),
        ],
        "law_refs": COMMON_LAW_REFS,
        "assumptions": [
            "1세대1주택 비과세 적용 (householdHouseCount=1, 보유 5년, 거주 5년, 비조정대상 취득).",
            "양도가액 10억 < 12억 → 전액 비과세, 단계 2 종료.",
            "다주택 중과 미적용 (1세대1주택).",
            "미등기 양도 아님. 감면세액 없음. 단독명의.",
        ],
        "issue_flags_text": [
            "IS_1SE_1HOUSE (info) — 1세대1주택 비과세 적용 확인",
            "RESIDENCE_MONTHS_USER_INPUT (info) — 사용자 입력 거주기간 사용",
            "NECESSARY_EXPENSE_BREAKDOWN_MISSING (info) — 필요경비 단일 입력",
            "UNREGISTERED_RATE_NOT_APPLIED (info) — 등기자산 가정",
            "ACQUISITION_CAUSE_ASSUMED_PURCHASE (info) — 매매취득 가정",
        ],
    },
    "TC-007": {
        "intent": "1세대1주택 + 12억 초과 (안분 + 표 2 64%)",
        "version": "v0.2",
        "inputs": [
            ("기준연도",           "baseYear",              2026,            "—",  ""),
            ("세대 구성원 수",     "householdMembers",      2,               "명", ""),
            ("기본공제 사용 여부", "basicDeductionUsed",    "false",         "—",  ""),
            ("1세대 보유 주택 수", "householdHouseCount",   1,               "개", ""),
            ("일시적 2주택 여부",  "isOneTimeTwoHouses",    "false",         "—",  ""),
            ("주택 ID",            None,                    "A",             "—",  ""),
            ("취득일",             "acquisitionDate",       "2018-06-15",    "—",  ""),
            ("취득가액",           "acquisitionPrice",      800_000_000,     "원", ""),
            ("필요경비",           "necessaryExpense",      30_000_000,      "원", ""),
            ("취득시 조정대상지역", "acquisitionRegulated", "false",         "—",  ""),
            ("거주기간",           "residenceMonths",       96,              "개월", "8년"),
            ("현재 거주 여부",     "livingNow",             "true",          "—",  ""),
            ("양도예정일",         "expectedSaleDate",      "2026-09-30",    "—",  ""),
            ("양도가액",           "expectedSalePrice",     1_500_000_000,   "원", "15억 (12억 초과)"),
            ("양도시 조정대상지역", "saleRegulated",        "false",         "—",  ""),
        ],
        "expected_steps": [
            ("transferGain",          "1,500,000,000 - 800,000,000 - 30,000,000",                                   670_000_000,   ""),
            ("is1Se1House",           "householdHouseCount=1, 보유 8.3년 ≥ 2년",                                    "true",        ""),
            ("isHighValueHouse",      "salePrice 15억 > 12억",                                                       "true",        "고가주택 안분 적용"),
            ("allocationRatio",       "(salePrice - 12억) / salePrice = (15억 - 12억) / 15억",                       0.20,          "안분비율"),
            ("taxableGain (안분 후)", "FLOOR(transferGain × allocationRatio)",                                      134_000_000,   ""),
            ("holdingYears",          "동월동일 기준 8년",                                                            8,             ""),
            ("residenceYears",        "FLOOR(residenceMonths / 12)",                                                8,             ""),
            ("longTermDeduction",     "FLOOR(taxableGain × (8×4% + 8×4%)) = FLOOR(134,000,000 × 0.64)",             85_760_000,    "표 2"),
            ("capitalGainIncome",     "taxableGain - longTermDeduction",                                            48_240_000,    ""),
            ("basicDeduction",        "2,500,000",                                                                  2_500_000,     ""),
            ("taxBase",               "capitalGainIncome - basicDeduction",                                         45_740_000,    ""),
            ("appliedRate (구간)",    "1,400만 ~ 5,000만 이하 (15%)",                                                 "1,400만~5,000만 이하 (15%)", ""),
            ("calculatedTax",         "840,000 + (45,740,000 - 14,000,000) × 0.15",                                 5_601_000,     "Claude 추정 (검증 필요)"),
            ("localIncomeTax",        "FLOOR(calculatedTax × 0.10)",                                                560_100,       ""),
            ("totalTax",              "calculatedTax + localIncomeTax",                                             6_161_100,     "최종"),
            ("netAfterTaxSaleAmount", "salePrice - totalTax",                                                       1_493_838_900, "최종"),
        ],
        "law_refs": COMMON_LAW_REFS,
        "assumptions": [
            "1세대1주택 + 12억 초과 → 12억 초과분만 과세 (고가주택 안분).",
            "안분비율 = (15억 - 12억) / 15억 = 20%.",
            "장특공 표 2: 보유 8년 × 4% + 거주 8년 × 4% = 64%.",
            "다주택 중과 미적용. 단독명의. 미등기 양도 아님.",
        ],
        "issue_flags_text": [
            "IS_1SE_1HOUSE (info)",
            "IS_HIGH_VALUE_HOUSE (info) — 안분비율 20%",
            "LONG_TERM_DEDUCTION_TABLE_2 (info) — 보유 32% + 거주 32% = 64%",
            "RESIDENCE_MONTHS_USER_INPUT (info)",
            "NECESSARY_EXPENSE_BREAKDOWN_MISSING (info)",
            "UNREGISTERED_RATE_NOT_APPLIED (info)",
            "ACQUISITION_CAUSE_ASSUMED_PURCHASE (info)",
        ],
    },
    "TC-008": {
        "intent": "다주택 일반과세 + 표 1 (보유 12년 → 24%)",
        "version": "v0.2",
        "inputs": [
            ("기준연도",           "baseYear",              2026,            "—",  ""),
            ("세대 구성원 수",     "householdMembers",      2,               "명", ""),
            ("기본공제 사용 여부", "basicDeductionUsed",    "false",         "—",  ""),
            ("1세대 보유 주택 수", "householdHouseCount",   2,               "개", "다주택 → 비과세 미적용"),
            ("일시적 2주택 여부",  "isOneTimeTwoHouses",    "false",         "—",  ""),
            ("주택 ID",            None,                    "A",             "—",  ""),
            ("취득일",             "acquisitionDate",       "2014-05-20",    "—",  ""),
            ("취득가액",           "acquisitionPrice",      500_000_000,     "원", ""),
            ("필요경비",           "necessaryExpense",      20_000_000,      "원", ""),
            ("취득시 조정대상지역", "acquisitionRegulated", "false",         "—",  ""),
            ("거주기간",           "residenceMonths",       0,               "개월", "거주 안 함"),
            ("현재 거주 여부",     "livingNow",             "false",         "—",  ""),
            ("양도예정일",         "expectedSaleDate",      "2026-08-15",    "—",  ""),
            ("양도가액",           "expectedSalePrice",     1_000_000_000,   "원", ""),
            ("양도시 조정대상지역", "saleRegulated",        "false",         "—",  "비조정 → 중과 미적용"),
        ],
        "expected_steps": [
            ("transferGain",          "1,000,000,000 - 500,000,000 - 20,000,000",                  480_000_000,   ""),
            ("is1Se1House",           "householdHouseCount=2",                                     "false",        "다주택"),
            ("isHighValueHouse",      "다주택 → N/A",                                              "—",           ""),
            ("taxableGain",           "안분 미적용. transferGain 그대로",                          480_000_000,   ""),
            ("holdingYears",          "2014-05-20 → 2026-08-15 = 12년",                            12,            ""),
            ("longTermDeduction",     "FLOOR(taxableGain × (0.06 + (12-3)×0.02)) = FLOOR(480,000,000 × 0.24)",   115_200_000,   "표 1"),
            ("capitalGainIncome",     "taxableGain - longTermDeduction",                          364_800_000,   ""),
            ("basicDeduction",        "2,500,000",                                                  2_500_000,     ""),
            ("taxBase",               "capitalGainIncome - basicDeduction",                       362_300_000,   ""),
            ("appliedRate (구간)",    "3억 ~ 5억 이하 (40%)",                                       "3억~5억 이하 (40%)", ""),
            ("calculatedTax",         "94,060,000 + (362,300,000 - 300,000,000) × 0.40",         118_980_000,   "Claude 추정"),
            ("localIncomeTax",        "FLOOR(calculatedTax × 0.10)",                              11_898_000,    ""),
            ("totalTax",              "calculatedTax + localIncomeTax",                          130_878_000,   "최종"),
            ("netAfterTaxSaleAmount", "salePrice - totalTax",                                    869_122_000,   "최종"),
        ],
        "law_refs": COMMON_LAW_REFS,
        "assumptions": [
            "다주택 (householdHouseCount=2) → 1세대1주택 비과세 미적용.",
            "비조정대상지역 양도 → 중과 미적용. 일반 누진세율.",
            "장특공 표 1 적용: 보유 12년 → 6% + (12-3)×2% = 24%.",
            "거주 안 함 → 표 2 자격 없음. 표 1만 적용.",
        ],
        "issue_flags_text": [
            "LONG_TERM_DEDUCTION_TABLE_1 (info) — 보유 12년 → 24%",
            "RESIDENCE_MONTHS_USER_INPUT (info)",
            "NECESSARY_EXPENSE_BREAKDOWN_MISSING (info)",
            "UNREGISTERED_RATE_NOT_APPLIED (info)",
            "ACQUISITION_CAUSE_ASSUMED_PURCHASE (info)",
        ],
    },
    "TC-009": {
        "intent": "1세대1주택 + 표 2 최대 80% (안분 + 14억)",
        "version": "v0.2",
        "inputs": [
            ("기준연도",           "baseYear",              2026,            "—",  ""),
            ("세대 구성원 수",     "householdMembers",      2,               "명", ""),
            ("기본공제 사용 여부", "basicDeductionUsed",    "false",         "—",  ""),
            ("1세대 보유 주택 수", "householdHouseCount",   1,               "개", ""),
            ("일시적 2주택 여부",  "isOneTimeTwoHouses",    "false",         "—",  ""),
            ("주택 ID",            None,                    "A",             "—",  ""),
            ("취득일",             "acquisitionDate",       "2016-04-30",    "—",  ""),
            ("취득가액",           "acquisitionPrice",      700_000_000,     "원", ""),
            ("필요경비",           "necessaryExpense",      25_000_000,      "원", ""),
            ("취득시 조정대상지역", "acquisitionRegulated", "false",         "—",  ""),
            ("거주기간",           "residenceMonths",       120,             "개월", "10년 (최대 클램프)"),
            ("현재 거주 여부",     "livingNow",             "true",          "—",  ""),
            ("양도예정일",         "expectedSaleDate",      "2026-09-15",    "—",  ""),
            ("양도가액",           "expectedSalePrice",     1_400_000_000,   "원", "14억 (12억 초과, 작업지시서 정정)"),
            ("양도시 조정대상지역", "saleRegulated",        "false",         "—",  ""),
        ],
        "expected_steps": [
            ("transferGain",          "1,400,000,000 - 700,000,000 - 25,000,000",                 675_000_000,   ""),
            ("is1Se1House",           "householdHouseCount=1, 보유 10년",                          "true",         ""),
            ("isHighValueHouse",      "salePrice 14억 > 12억",                                     "true",         ""),
            ("allocationRatio",       "(14억 - 12억) / 14억 = 1/7 ≈ 0.142857",                     "1/7",          ""),
            ("taxableGain (안분 후)", "FLOOR(675,000,000 × 1/7) = FLOOR(96,428,571.428…)",       96_428_571,    ""),
            ("holdingYears",          "10년",                                                       10,            ""),
            ("residenceYears",        "10년 (클램프)",                                              10,            ""),
            ("longTermDeduction",     "FLOOR(96,428,571 × 0.80) — 보유 40% + 거주 40% = 80% 최대", 77_142_856,    "표 2 최대"),
            ("capitalGainIncome",     "taxableGain - longTermDeduction",                          19_285_715,    ""),
            ("basicDeduction",        "2,500,000",                                                  2_500_000,     ""),
            ("taxBase",               "capitalGainIncome - basicDeduction",                       16_785_715,    ""),
            ("appliedRate (구간)",    "1,400만 ~ 5,000만 이하 (15%)",                              "1,400만~5,000만 이하 (15%)", ""),
            ("calculatedTax",         "840,000 + (16,785,715 - 14,000,000) × 0.15",                1_257_857,     "Claude 추정"),
            ("localIncomeTax",        "FLOOR(calculatedTax × 0.10)",                              125_785,       ""),
            ("totalTax",              "calculatedTax + localIncomeTax",                          1_383_642,     "최종"),
            ("netAfterTaxSaleAmount", "salePrice - totalTax",                                    1_398_616_358, "최종"),
        ],
        "law_refs": COMMON_LAW_REFS,
        "assumptions": [
            "1세대1주택 + 12억 초과 → 안분 + 표 2 최대 80% 동시 검증.",
            "원안 11억은 12억 이하라 표 2가 호출되지 않아 검증 불가능 → 14억으로 변경.",
            "장특공 표 2 최대: 보유 10년 → 40% + 거주 10년 → 40% = 80%.",
            "안분비율 1/7 ≈ 14.29%.",
        ],
        "issue_flags_text": [
            "IS_1SE_1HOUSE (info)",
            "IS_HIGH_VALUE_HOUSE (info) — 안분비율 14.29%",
            "LONG_TERM_DEDUCTION_TABLE_2 (info) — 보유 40% + 거주 40% = 80% 최대",
            "RESIDENCE_MONTHS_USER_INPUT (info)",
            "NECESSARY_EXPENSE_BREAKDOWN_MISSING (info)",
            "UNREGISTERED_RATE_NOT_APPLIED (info)",
            "ACQUISITION_CAUSE_ASSUMED_PURCHASE (info)",
        ],
    },
    "TC-010": {
        "intent": "일시적 2주택 (적용 안 함, 다주택 일반과세)",
        "version": "v0.2",
        "inputs": [
            ("기준연도",           "baseYear",              2026,            "—",  ""),
            ("세대 구성원 수",     "householdMembers",      2,               "명", ""),
            ("기본공제 사용 여부", "basicDeductionUsed",    "false",         "—",  ""),
            ("1세대 보유 주택 수", "householdHouseCount",   2,               "개", "다주택"),
            ("일시적 2주택 여부",  "isOneTimeTwoHouses",    "true",          "—",  "issueFlag만 발동"),
            ("주택 ID",            None,                    "A",             "—",  ""),
            ("취득일",             "acquisitionDate",       "2021-05-20",    "—",  ""),
            ("취득가액",           "acquisitionPrice",      600_000_000,     "원", ""),
            ("필요경비",           "necessaryExpense",      15_000_000,      "원", ""),
            ("취득시 조정대상지역", "acquisitionRegulated", "false",         "—",  ""),
            ("거주기간",           "residenceMonths",       0,               "개월", ""),
            ("현재 거주 여부",     "livingNow",             "false",         "—",  ""),
            ("양도예정일",         "expectedSaleDate",      "2026-08-31",    "—",  ""),
            ("양도가액",           "expectedSalePrice",     1_000_000_000,   "원", ""),
            ("양도시 조정대상지역", "saleRegulated",        "false",         "—",  ""),
        ],
        "expected_steps": [
            ("transferGain",          "1,000,000,000 - 600,000,000 - 15,000,000",                 385_000_000,   ""),
            ("is1Se1House",           "householdHouseCount=2",                                     "false",        "다주택"),
            ("ONE_TIME_2HOUSES_NOT_APPLIED", "issueFlag warning 발동만, 산식 분기 없음",            "warning",     ""),
            ("taxableGain",           "안분 미적용",                                                385_000_000,   ""),
            ("holdingYears",          "2021-05-20 → 2026-08-31 = 5년",                            5,             ""),
            ("longTermDeduction",     "FLOOR(transferGain × (0.06 + (5-3)×0.02)) = FLOOR(385,000,000 × 0.10)",  38_500_000,    "표 1"),
            ("capitalGainIncome",     "taxableGain - longTermDeduction",                          346_500_000,   ""),
            ("basicDeduction",        "2,500,000",                                                  2_500_000,     ""),
            ("taxBase",               "capitalGainIncome - basicDeduction",                       344_000_000,   ""),
            ("appliedRate (구간)",    "3억 ~ 5억 이하 (40%)",                                       "3억~5억 이하 (40%)", ""),
            ("calculatedTax",         "94,060,000 + (344,000,000 - 300,000,000) × 0.40",         111_660_000,   "Claude 추정"),
            ("localIncomeTax",        "FLOOR(calculatedTax × 0.10)",                              11_166_000,    ""),
            ("totalTax",              "calculatedTax + localIncomeTax",                          122_826_000,   "최종"),
            ("netAfterTaxSaleAmount", "salePrice - totalTax",                                    877_174_000,   "최종"),
        ],
        "law_refs": COMMON_LAW_REFS,
        "assumptions": [
            "isOneTimeTwoHouses=true 입력 → ONE_TIME_2HOUSES_NOT_APPLIED warning 발동.",
            "v0.2는 시행령 제155조 ① 본격 처리 미적용. 다주택 일반과세 그대로.",
            "v0.3 또는 v0.5에서 본격 처리 시 본 케이스의 totalTax는 0으로 변경 가능.",
            "장특공 표 1: 보유 5년 → 6% + (5-3)×2% = 10%.",
        ],
        "issue_flags_text": [
            "ONE_TIME_2HOUSES_NOT_APPLIED (warning) — v0.2 미적용",
            "LONG_TERM_DEDUCTION_TABLE_1 (info) — 보유 5년 → 10%",
            "RESIDENCE_MONTHS_USER_INPUT (info)",
            "NECESSARY_EXPENSE_BREAKDOWN_MISSING (info)",
            "UNREGISTERED_RATE_NOT_APPLIED (info)",
            "ACQUISITION_CAUSE_ASSUMED_PURCHASE (info)",
        ],
    },
}

# ============================================================
# 메인
# ============================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)

    make_rules_sheet(wb)

    for i, tc_id in enumerate(["TC-006", "TC-007", "TC-008", "TC-009", "TC-010"], start=1):
        d = TC_DATA[tc_id]
        sheet_name = f"{i}_{tc_id}"
        make_tc_sheet(wb, sheet_name, tc_id, d["intent"], d["version"],
                       d["inputs"], d["expected_steps"], d["law_refs"],
                       d["assumptions"], d["issue_flags_text"])

    make_rate_sheet(wb)
    make_deduction_sheet(wb)

    output_dir = "docs/v0.2"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "04_test_cases_manual.xlsx")
    wb.save(output_path)

    print(f"OK: {output_path}")
    print(f"sheets: {len(wb.sheetnames)}")
    for sn in wb.sheetnames:
        print(f"  - {sn}")

if __name__ == "__main__":
    main()

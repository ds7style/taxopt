# -*- coding: utf-8 -*-
"""
TaxOpt v0.3-A 검증팀 손계산 양식 생성 스크립트
v0.2 generate_v02_xlsx.py 패턴 그대로 계승 + 다주택 중과 영역 추가.

시트 구조 (8 시트):
  0_규칙           — 작성 규칙·표준 변수명·색상 범례 + v0.3-A 신규 변수 명시
  1_TC-011         — 2주택 + saleRegulated=true (중과 +20%p + 장특공 배제)
  2_TC-012         — 3주택 + saleRegulated=true (중과 +30%p)
  3_TC-013         — 2주택 + saleRegulated=false 회귀 (TC-008과 동일)
  4_TC-014         — 3주택 + saleRegulated=false 회귀 (TC-008과 동일)
  5_세율표         — 8단계 누진세율표 + 가산세율 룩업 (v0.3-A 신규)
  6_장특공표       — 표 1·표 2 (v0.2 그대로)
  7_중과판정       — 4단계 조건 흐름도 + issueFlag 카탈로그

색상 범례 (v0.2 양식 그대로):
  파란 글자 (FF0000FF): 입력값
  검정 글자 (FF000000): 수식·계산 결과
  연두 배경 (FFE2EFDA): 최종 결과 셀
  진한 파랑 배경 (FF1F4E78) + 흰 글자: 시트 제목
  중간 파랑 배경 (FF2E75B6) + 흰 글자: 섹션 헤더
  연한 파랑 배경 (FFDEEBF7): 라벨 셀
  주황 배경 (FFFFD966): v0.3-A 신규 영역 (중과 관련)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 스타일 (v0.2 양식 그대로 + 중과 영역 신규)
# ============================================================
TITLE_FILL    = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FILL   = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
LABEL_FILL    = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
RESULT_FILL   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HEAVY_FILL    = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # v0.3-A 신규

WHITE_BOLD    = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
BLACK_BOLD    = Font(name="맑은 고딕", size=11, bold=True, color="000000")
BLACK_NORMAL  = Font(name="맑은 고딕", size=11, bold=False, color="000000")
BLUE_INPUT    = Font(name="맑은 고딕", size=11, bold=False, color="0000FF")
TITLE_FONT    = Font(name="맑은 고딕", size=13, bold=True, color="FFFFFF")

THIN          = Side(style='thin', color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT          = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT         = Alignment(horizontal='right',  vertical='center', wrap_text=False)


def setup_columns(ws, widths):
    """열 너비 설정"""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def title_row(ws, row, text, span=4):
    """시트 제목 행"""
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    ws.cell(row=row, column=1).fill = TITLE_FILL
    ws.cell(row=row, column=1).alignment = CENTER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 28


def section_header(ws, row, text, span=4):
    """섹션 헤더"""
    ws.cell(row=row, column=1, value=text).font = WHITE_BOLD
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22


def label_cell(ws, row, col, text):
    """라벨 셀 (연한 파랑)"""
    c = ws.cell(row=row, column=col, value=text)
    c.font = BLACK_BOLD
    c.fill = LABEL_FILL
    c.alignment = LEFT
    c.border = BORDER


def input_cell(ws, row, col, value, fmt=None):
    """입력 셀 (파란 글자)"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = BLUE_INPUT
    c.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def calc_cell(ws, row, col, value, fmt=None, bold=False):
    """계산 셀 (검정 글자)"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = BLACK_BOLD if bold else BLACK_NORMAL
    c.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def result_cell(ws, row, col, value, fmt=None):
    """최종 결과 셀 (연두 배경)"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = BLACK_BOLD
    c.fill = RESULT_FILL
    c.alignment = RIGHT
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def heavy_cell(ws, row, col, value, fmt=None):
    """v0.3-A 중과 영역 셀 (주황 배경)"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = BLACK_BOLD
    c.fill = HEAVY_FILL
    c.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
    c.border = BORDER
    if fmt:
        c.number_format = fmt


# ============================================================
# 시트 0: 규칙
# ============================================================
def build_sheet_rules(wb):
    ws = wb.create_sheet("0_규칙")
    setup_columns(ws, [22, 22, 22, 50])

    title_row(ws, 1, "TaxOpt v0.3-A 검증팀 손계산 양식 — 작성 규칙", span=4)

    section_header(ws, 3, "1. 양식 정보", span=4)
    label_cell(ws, 4, 1, "버전")
    calc_cell(ws, 4, 2, "v0.3-A")
    label_cell(ws, 4, 3, "베이스")
    calc_cell(ws, 4, 4, "v0.2 양식 패턴 계승 + 다주택 중과 영역")

    label_cell(ws, 5, 1, "범위")
    calc_cell(ws, 5, 2, "TC-011~014 4건")
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)

    label_cell(ws, 6, 1, "검증 대상")
    calc_cell(ws, 6, 2, "검증팀 손계산 + 홈택스 모의계산")
    label_cell(ws, 6, 3, "최종")
    calc_cell(ws, 6, 4, "Claude 명세서·코드와 5자 일치 확인")

    section_header(ws, 8, "2. 색상 범례", span=4)
    label_cell(ws, 9, 1, "파란 글자")
    input_cell(ws, 9, 2, "예: 1,000,000,000")
    calc_cell(ws, 9, 3, "입력값")
    calc_cell(ws, 9, 4, "작성자가 직접 채움")

    label_cell(ws, 10, 1, "검정 글자")
    calc_cell(ws, 10, 2, "예: 480,000,000")
    calc_cell(ws, 10, 3, "수식·계산")
    calc_cell(ws, 10, 4, "라벨 또는 산출 결과")

    label_cell(ws, 11, 1, "연두 배경")
    result_cell(ws, 11, 2, "286,616,000")
    calc_cell(ws, 11, 3, "최종 결과")
    calc_cell(ws, 11, 4, "totalTax 등 핵심 결과")

    label_cell(ws, 12, 1, "주황 배경")
    heavy_cell(ws, 12, 2, "중과 영역")
    calc_cell(ws, 12, 3, "v0.3-A 신규")
    calc_cell(ws, 12, 4, "다주택 중과 관련 신규 변수·산식")

    section_header(ws, 14, "3. 표준 변수명 (v0.3-A 신규)", span=4)
    headers = ["변수명", "타입", "단계", "의미"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=15, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    rules_data = [
        ("isHeavyTaxation", "boolean", "4", "다주택 중과 발동 여부 (4단계 조건 AND)"),
        ("heavyRateAddition", "number", "4", "가산세율 (0.20=2주택, 0.30=3주택+)"),
        ("appliedDeductionTable", "1|2|null", "4", "장특공 표 (중과 시 null)"),
        ("longTermDeduction", "number", "4", "장특공 (중과 시 0)"),
        ("appliedRate.addition", "number", "9", "세율 가산분 (heavyRateAddition 인용)"),
        ("baseTax_with_addition", "number", "9", "구간별 누적 baseTax + 가산세율"),
    ]
    for i, (var, typ, step, desc) in enumerate(rules_data, start=16):
        calc_cell(ws, i, 1, var)
        calc_cell(ws, i, 2, typ)
        calc_cell(ws, i, 3, step)
        calc_cell(ws, i, 4, desc)

    section_header(ws, 23, "4. 4단계 조건 (다주택 중과 판정)", span=4)
    cond_data = [
        ("조건 1", "다주택 (householdHouseCount ≥ 2)", "v0.2 보존 필드"),
        ("조건 2", "양도시 조정대상 (saleRegulated = true)", "v0.2 보존 → v0.3-A 활성"),
        ("조건 3", "양도일 ≥ 2026-05-10", "의사결정 #1 (중과 유예 종료)"),
        ("조건 4", "비과세 미적용 (is1Se1House = false)", "v0.2 단계 2 결과"),
    ]
    for i, (cond, body, ref) in enumerate(cond_data, start=24):
        label_cell(ws, i, 1, cond)
        calc_cell(ws, i, 2, body)
        calc_cell(ws, i, 3, ref)
        calc_cell(ws, i, 4, "")
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)

    label_cell(ws, 28, 1, "AND 결과")
    heavy_cell(ws, 28, 2, "isHeavyTaxation = 조건1 ∧ 조건2 ∧ 조건3 ∧ 조건4")
    ws.merge_cells(start_row=28, start_column=2, end_row=28, end_column=4)


# ============================================================
# 공통 — TC 시트 빌더
# ============================================================
def build_tc_sheet(wb, tc_id, sheet_idx, title_text, inputs, steps, total_tax, net_after_tax,
                   issue_flags, validation_intent, is_heavy):
    ws = wb.create_sheet(f"{sheet_idx}_{tc_id}")
    setup_columns(ws, [25, 22, 18, 50])

    title_row(ws, 1, f"{tc_id} — {title_text}", span=4)

    # 입력
    section_header(ws, 3, "입력 (caseData)", span=4)
    headers = ["필드", "값", "타입", "비고"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    row = 5
    for field, value, typ, note in inputs:
        label_cell(ws, row, 1, field)
        if typ == "highlight":
            heavy_cell(ws, row, 2, value)
        else:
            input_cell(ws, row, 2, value)
        calc_cell(ws, row, 3, typ if typ != "highlight" else "변경 핵심")
        calc_cell(ws, row, 4, note)
        row += 1

    # 단계별 기대값
    row += 1
    section_header(ws, row, "단계별 기대값 (검증팀 작성 영역)", span=4)
    row += 1
    headers2 = ["단계", "변수", "기대값", "산식 / 비고"]
    for i, h in enumerate(headers2, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    row += 1

    for step, var, expected, formula, is_highlight in steps:
        calc_cell(ws, row, 1, step)
        calc_cell(ws, row, 2, var)
        if is_highlight:
            heavy_cell(ws, row, 3, expected)
        else:
            input_cell(ws, row, 3, expected)
        calc_cell(ws, row, 4, formula)
        row += 1

    # 최종 결과
    row += 1
    section_header(ws, row, "최종 결과", span=4)
    row += 1
    label_cell(ws, row, 1, "totalTax")
    result_cell(ws, row, 2, total_tax, fmt="#,##0")
    calc_cell(ws, row, 3, "원")
    calc_cell(ws, row, 4, "12단계 — 검증 대기" if is_heavy else "12단계 — 회귀")
    row += 1
    label_cell(ws, row, 1, "netAfterTaxSaleAmount")
    result_cell(ws, row, 2, net_after_tax, fmt="#,##0")
    calc_cell(ws, row, 3, "원")
    calc_cell(ws, row, 4, "13단계 — 세후 매각금액")

    # issueFlag
    row += 2
    section_header(ws, row, "issueFlag (기대 발동)", span=4)
    row += 1
    headers3 = ["code", "severity", "발동", "비고"]
    for i, h in enumerate(headers3, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    row += 1
    for code, severity, fired, note in issue_flags:
        calc_cell(ws, row, 1, code)
        calc_cell(ws, row, 2, severity)
        if fired == "✅":
            heavy_cell(ws, row, 3, fired)
        else:
            calc_cell(ws, row, 3, fired)
        calc_cell(ws, row, 4, note)
        row += 1

    # 검증 의도
    row += 1
    section_header(ws, row, "검증 의도", span=4)
    row += 1
    for line in validation_intent:
        calc_cell(ws, row, 1, line)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1


# ============================================================
# 시트 1: TC-011
# ============================================================
def build_sheet_tc011(wb):
    inputs = [
        ("baseYear", 2026, "number", ""),
        ("basicDeductionUsed", "FALSE", "boolean", ""),
        ("householdHouseCount", 2, "number", "다주택"),
        ("isOneTimeTwoHouses", "FALSE", "boolean", ""),
        ("houses[0].id", "A", "string", ""),
        ("houses[0].acquisitionDate", "2014-05-20", "date", "TC-008과 동일 (보유 12년)"),
        ("houses[0].acquisitionPrice", 500000000, "number", "TC-008과 동일"),
        ("houses[0].necessaryExpense", 20000000, "number", "TC-008과 동일"),
        ("houses[0].acquisitionRegulated", "FALSE", "boolean", ""),
        ("houses[0].residenceMonths", 0, "number", "거주 안 함"),
        ("houses[0].livingNow", "FALSE", "boolean", ""),
        ("houses[0].expectedSaleDate", "2026-08-15", "date", "중과 유예 종료 후"),
        ("houses[0].expectedSalePrice", 1000000000, "number", "TC-008과 동일 (10억)"),
        ("houses[0].saleRegulated", "TRUE", "highlight", "양도시 조정대상지역 → 중과 발동"),
    ]
    steps = [
        (1, "transferGain", 480000000, "1,000,000,000 − 500,000,000 − 20,000,000", False),
        (2, "is1Se1House", "FALSE", "householdHouseCount=2", False),
        (2, "taxableGain", 480000000, "passthrough (다주택 → 비과세 미적용)", False),
        (3, "taxableGain", 480000000, "안분 미적용", False),
        (4, "isHeavyTaxation", "TRUE", "4단계 조건 모두 충족 ✅", True),
        (4, "heavyRateAddition", 0.20, "2주택 → +20%p", True),
        (4, "holdingYears", 12, "2014-05-20 → 2026-08-15", False),
        (4, "appliedDeductionTable", "null", "중과 배제 (제95조 ② 단서)", True),
        (4, "longTermDeduction", 0, "중과 → 표 1·2 모두 배제", True),
        (5, "capitalGainIncome", 480000000, "480,000,000 − 0", False),
        (6, "basicDeduction", 2500000, "basicDeductionUsed=false", False),
        (7, "taxBase", 477500000, "480,000,000 − 2,500,000", False),
        (8, "holdingPeriodBranch", "over2y", "보유 12년", False),
        (9, "appliedRate.bracket", "3억 ~ 5억 (40%)", "findBracket(477,500,000)", False),
        (9, "appliedRate.addition", 0.20, "가산세율 (heavyRateAddition 인용)", True),
        (9, "baseTax (기본)", 94060000, "누진세율표 그대로", False),
        (9, "baseTax_with_addition", 154060000, "3억 lowerBound까지 누적 (+20%p)", True),
        (9, "calculatedTax", 260560000, "154,060,000 + (477,500,000−300,000,000)×0.60", True),
        (10, "calculatedTax (최종)", 260560000, "10단계 산출세액", False),
        (11, "localIncomeTax", 26056000, "floor(260,560,000 × 0.1)", False),
    ]
    issue_flags = [
        ("HEAVY_TAXATION_APPLIED", "warning", "✅", "중과 발동"),
        ("HEAVY_TAXATION_2_HOUSES", "info", "✅", "2주택 분기"),
        ("LONG_TERM_DEDUCTION_EXCLUDED_BY_MULTI_HOUSE_HEAVY", "info", "✅", "보유 12년 무관 배제"),
        ("SALE_REGULATED_USER_INPUT", "info", "✅", "사용자 직접 입력 (B-033)"),
        ("HEAVY_TAX_EXCLUSION_NOT_HANDLED", "info", "✅", "시행령 제167조의10·11 단서 미처리"),
        ("RESIDENCE_MONTHS_USER_INPUT", "info", "✅", ""),
        ("NECESSARY_EXPENSE_BREAKDOWN_MISSING", "info", "✅", ""),
        ("UNREGISTERED_RATE_NOT_APPLIED", "info", "✅", ""),
        ("ACQUISITION_CAUSE_ASSUMED_PURCHASE", "info", "✅", ""),
        ("LONG_TERM_DEDUCTION_TABLE_1", "info", "❌", "중과 배제로 미발동"),
    ]
    intent = [
        "▪ TC-008과 비교: saleRegulated만 false → true 변경. 중과 +20%p 발동 + 장특공 배제.",
        "▪ totalTax 130,878,000 (TC-008) → 286,616,000 (TC-011), 약 2.19배 증가.",
        "▪ 산식 검증: 누진 구간 누적 baseTax_with_addition + (taxBase − lowerBound) × (marginalRate + addition).",
        "▪ 검증팀 손계산 + 홈택스 모의계산 (홈택스는 다주택 중과 분기 자동 처리).",
    ]
    build_tc_sheet(
        wb, "TC-011", 1,
        "2주택 + saleRegulated=true (중과 +20%p + 장특공 배제)",
        inputs, steps,
        total_tax=286616000, net_after_tax=713384000,
        issue_flags=issue_flags, validation_intent=intent, is_heavy=True
    )


# ============================================================
# 시트 2: TC-012
# ============================================================
def build_sheet_tc012(wb):
    inputs = [
        ("baseYear", 2026, "number", ""),
        ("basicDeductionUsed", "FALSE", "boolean", ""),
        ("householdHouseCount", 3, "highlight", "3주택 (중과 +30%p)"),
        ("isOneTimeTwoHouses", "FALSE", "boolean", ""),
        ("houses[0].id", "A", "string", ""),
        ("houses[0].acquisitionDate", "2014-05-20", "date", "TC-011과 동일"),
        ("houses[0].acquisitionPrice", 500000000, "number", "TC-011과 동일"),
        ("houses[0].necessaryExpense", 20000000, "number", "TC-011과 동일"),
        ("houses[0].acquisitionRegulated", "FALSE", "boolean", ""),
        ("houses[0].residenceMonths", 0, "number", ""),
        ("houses[0].livingNow", "FALSE", "boolean", ""),
        ("houses[0].expectedSaleDate", "2026-08-15", "date", ""),
        ("houses[0].expectedSalePrice", 1000000000, "number", ""),
        ("houses[0].saleRegulated", "TRUE", "highlight", ""),
    ]
    steps = [
        (1, "transferGain", 480000000, "(TC-011과 동일)", False),
        (2, "is1Se1House", "FALSE", "", False),
        (3, "taxableGain", 480000000, "", False),
        (4, "isHeavyTaxation", "TRUE", "", True),
        (4, "heavyRateAddition", 0.30, "3주택 클램프 → +30%p", True),
        (4, "appliedDeductionTable", "null", "중과 배제", True),
        (4, "longTermDeduction", 0, "", True),
        (5, "capitalGainIncome", 480000000, "", False),
        (6, "basicDeduction", 2500000, "", False),
        (7, "taxBase", 477500000, "", False),
        (8, "holdingPeriodBranch", "over2y", "", False),
        (9, "appliedRate.bracket", "3억 ~ 5억 (40%)", "", False),
        (9, "appliedRate.addition", 0.30, "가산세율", True),
        (9, "baseTax_with_addition", 184060000, "3억 lowerBound까지 (+30%p 누적)", True),
        (9, "calculatedTax", 308310000, "184,060,000 + (477,500,000−300,000,000)×0.70", True),
        (11, "localIncomeTax", 30831000, "floor(308,310,000 × 0.1)", False),
    ]
    issue_flags = [
        ("HEAVY_TAXATION_APPLIED", "warning", "✅", "중과 발동"),
        ("HEAVY_TAXATION_3_HOUSES", "info", "✅", "3주택 분기"),
        ("HEAVY_TAXATION_2_HOUSES", "info", "❌", "3주택 분기로 미발동"),
        ("LONG_TERM_DEDUCTION_EXCLUDED_BY_MULTI_HOUSE_HEAVY", "info", "✅", ""),
        ("SALE_REGULATED_USER_INPUT", "info", "✅", ""),
        ("HEAVY_TAX_EXCLUSION_NOT_HANDLED", "info", "✅", ""),
        ("(보조 4종)", "info", "✅", "TC-011과 동일"),
    ]
    intent = [
        "▪ TC-011 (+20%p) vs TC-012 (+30%p) 비교: totalTax 286,616,000 → 339,141,000.",
        "▪ 가산세율 룩업 클램프 검증: householdHouseCount=3 → findHeavyTaxRateAddition → addition=0.30.",
        "▪ 약 +52,525,000 증가.",
        "▪ 검증팀 손계산 + 홈택스 모의계산.",
    ]
    build_tc_sheet(
        wb, "TC-012", 2,
        "3주택 + saleRegulated=true (중과 +30%p)",
        inputs, steps,
        total_tax=339141000, net_after_tax=660859000,
        issue_flags=issue_flags, validation_intent=intent, is_heavy=True
    )


# ============================================================
# 시트 3: TC-013 (회귀)
# ============================================================
def build_sheet_tc013(wb):
    inputs = [
        ("baseYear", 2026, "number", ""),
        ("basicDeductionUsed", "FALSE", "boolean", ""),
        ("householdHouseCount", 2, "number", ""),
        ("isOneTimeTwoHouses", "FALSE", "boolean", ""),
        ("houses[0].acquisitionDate", "2014-05-20", "date", ""),
        ("houses[0].acquisitionPrice", 500000000, "number", ""),
        ("houses[0].necessaryExpense", 20000000, "number", ""),
        ("houses[0].acquisitionRegulated", "FALSE", "boolean", ""),
        ("houses[0].residenceMonths", 0, "number", ""),
        ("houses[0].expectedSaleDate", "2026-08-15", "date", ""),
        ("houses[0].expectedSalePrice", 1000000000, "number", ""),
        ("houses[0].saleRegulated", "FALSE", "highlight", "TC-011과 다름. 중과 미발동."),
    ]
    steps = [
        (4, "isHeavyTaxation", "FALSE", "조건 2 미충족 (saleRegulated=false)", False),
        (4, "heavyRateAddition", "null", "중과 미발동", False),
        (4, "appliedDeductionTable", 1, "표 1 적용 (보유 12년 → 24%)", False),
        (4, "longTermDeduction", 115200000, "TC-008과 동일", False),
        (9, "appliedRate.type", "progressive", "중과 가산 없음", False),
        (9, "calculatedTax", 118980000, "TC-008과 동일", False),
        (11, "localIncomeTax", 11898000, "TC-008과 동일", False),
    ]
    issue_flags = [
        ("HEAVY_TAXATION_APPLIED", "warning", "❌", "조건 2 미충족"),
        ("LONG_TERM_DEDUCTION_TABLE_1", "info", "✅", "표 1 적용 (보유 12년)"),
        ("SALE_REGULATED_USER_INPUT", "info", "✅", ""),
        ("(보조 4종)", "info", "✅", "TC-008과 동일"),
    ]
    intent = [
        "▪ v0.2 회귀 안전성 검증 (가장 중요): saleRegulated=false 케이스에서 v0.3-A 코드가 v0.2 결과를 그대로 보존하는지 확인.",
        "▪ 중과 4단계 조건 중 조건 2 단독으로 차단되는 케이스.",
        "▪ TC-008은 v0.3-A 명세서에서 TC-013으로 재명명. 골든셋 일람표에 v0.2 회귀 케이스로 분류.",
        "▪ 본 케이스가 깨지면 v0.2 → v0.3-A 마이그레이션 실패 → 즉시 롤백.",
    ]
    build_tc_sheet(
        wb, "TC-013", 3,
        "2주택 + saleRegulated=false (중과 미발동, v0.2 회귀)",
        inputs, steps,
        total_tax=130878000, net_after_tax=869122000,
        issue_flags=issue_flags, validation_intent=intent, is_heavy=False
    )


# ============================================================
# 시트 4: TC-014 (회귀, 보강)
# ============================================================
def build_sheet_tc014(wb):
    inputs = [
        ("baseYear", 2026, "number", ""),
        ("basicDeductionUsed", "FALSE", "boolean", ""),
        ("householdHouseCount", 3, "highlight", "3주택"),
        ("isOneTimeTwoHouses", "FALSE", "boolean", ""),
        ("houses[0].acquisitionDate", "2014-05-20", "date", ""),
        ("houses[0].acquisitionPrice", 500000000, "number", ""),
        ("houses[0].necessaryExpense", 20000000, "number", ""),
        ("houses[0].acquisitionRegulated", "FALSE", "boolean", ""),
        ("houses[0].residenceMonths", 0, "number", ""),
        ("houses[0].expectedSaleDate", "2026-08-15", "date", ""),
        ("houses[0].expectedSalePrice", 1000000000, "number", ""),
        ("houses[0].saleRegulated", "FALSE", "highlight", "중과 미발동 (조건 2 미충족)"),
    ]
    steps = [
        (4, "isHeavyTaxation", "FALSE", "조건 2 미충족", False),
        (4, "appliedDeductionTable", 1, "다주택 + 보유 12년 → 표 1 (24%)", False),
        (4, "longTermDeduction", 115200000, "TC-008과 동일", False),
        (12, "totalTax", 130878000, "TC-008·013과 동일 (saleRegulated=false 핵심)", False),
    ]
    issue_flags = [
        ("HEAVY_TAXATION_APPLIED", "warning", "❌", "조건 2 미충족"),
        ("LONG_TERM_DEDUCTION_TABLE_1", "info", "✅", ""),
        ("SALE_REGULATED_USER_INPUT", "info", "✅", ""),
    ]
    intent = [
        "▪ 3주택 + 비조정대상 회귀 케이스: 다주택 보유 자체가 아닌 saleRegulated가 중과의 키임을 검증.",
        "▪ householdHouseCount=3이라도 saleRegulated=false이면 일반과세 (표 1 적용).",
        "▪ TC-008 = TC-013 = TC-014 = 130,878,000 (3건 모두 동일).",
        "▪ 검증팀 손계산 + 홈택스 모의계산.",
    ]
    build_tc_sheet(
        wb, "TC-014", 4,
        "3주택 + saleRegulated=false (회귀, 보강)",
        inputs, steps,
        total_tax=130878000, net_after_tax=869122000,
        issue_flags=issue_flags, validation_intent=intent, is_heavy=False
    )


# ============================================================
# 시트 5: 세율표 + 가산세율
# ============================================================
def build_sheet_rates(wb):
    ws = wb.create_sheet("5_세율표")
    setup_columns(ws, [22, 22, 18, 50])

    title_row(ws, 1, "8단계 누진세율표 + 가산세율 룩업 (v0.3-A 신규)", span=4)

    section_header(ws, 3, "1. 누진세율표 (소득세법 제55조 ① — v0.1·v0.2·v0.3-A 동일)", span=4)
    headers = ["과세표준 구간", "세율", "lowerBound", "baseTax (lowerBound 직전 누적)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    brackets = [
        ("~ 14,000,000", 0.06, 0, 0),
        ("14,000,001 ~ 50,000,000", 0.15, 14000000, 840000),
        ("50,000,001 ~ 88,000,000", 0.24, 50000000, 6240000),
        ("88,000,001 ~ 150,000,000", 0.35, 88000000, 15360000),
        ("150,000,001 ~ 300,000,000", 0.38, 150000000, 37060000),
        ("300,000,001 ~ 500,000,000", 0.40, 300000000, 94060000),
        ("500,000,001 ~ 1,000,000,000", 0.42, 500000000, 174060000),
        ("1,000,000,001 ~", 0.45, 1000000000, 384060000),
    ]
    for i, (rng, rate, lower, base) in enumerate(brackets, start=5):
        calc_cell(ws, i, 1, rng)
        calc_cell(ws, i, 2, rate, fmt="0.0%")
        calc_cell(ws, i, 3, lower, fmt="#,##0")
        calc_cell(ws, i, 4, base, fmt="#,##0")

    section_header(ws, 14, "2. 가산세율 룩업 HEAVY_TAX_RATE_ADDITION (v0.3-A 신규)", span=4)
    headers2 = ["주택 수", "가산세율", "법령", "비고"]
    for i, h in enumerate(headers2, start=1):
        c = ws.cell(row=15, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    heavy_data = [
        ("2주택", 0.20, "제104조 ⑦", "+20%p"),
        ("3주택+", 0.30, "제104조 ⑦", "+30%p (3주택 이상 클램프)"),
    ]
    for i, (cnt, rate, law, note) in enumerate(heavy_data, start=16):
        label_cell(ws, i, 1, cnt)
        heavy_cell(ws, i, 2, rate, fmt="0.0%")
        calc_cell(ws, i, 3, law)
        calc_cell(ws, i, 4, note)

    section_header(ws, 19, "3. 누적 baseTax_with_addition (TC-011·012 산출 참고)", span=4)
    headers3 = ["과세표준 lowerBound", "baseTax (기본)", "+20%p 누적", "+30%p 누적"]
    for i, h in enumerate(headers3, start=1):
        c = ws.cell(row=20, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # +20%p 누적 = base + (각 구간 길이 × addition 누적)
    # 명세서 §3-4-3 표 본문에서 산출
    cum_data = [
        (0, 0, 0, 0),
        (14000000, 840000, 840000 + 14000000*0.20, 840000 + 14000000*0.30),
        (50000000, 6240000, 6240000 + 50000000*0.20, 6240000 + 50000000*0.30),
        (88000000, 15360000, 15360000 + 88000000*0.20, 15360000 + 88000000*0.30),
        (150000000, 37060000, 37060000 + 150000000*0.20, 37060000 + 150000000*0.30),
        (300000000, 94060000, 94060000 + 300000000*0.20, 94060000 + 300000000*0.30),
        (500000000, 174060000, 174060000 + 500000000*0.20, 174060000 + 500000000*0.30),
        (1000000000, 384060000, 384060000 + 1000000000*0.20, 384060000 + 1000000000*0.30),
    ]
    for i, (lower, base, plus20, plus30) in enumerate(cum_data, start=21):
        calc_cell(ws, i, 1, lower, fmt="#,##0")
        calc_cell(ws, i, 2, base, fmt="#,##0")
        heavy_cell(ws, i, 3, int(plus20), fmt="#,##0")
        heavy_cell(ws, i, 4, int(plus30), fmt="#,##0")


# ============================================================
# 시트 6: 장특공표 (v0.2 그대로)
# ============================================================
def build_sheet_long_term(wb):
    ws = wb.create_sheet("6_장특공표")
    setup_columns(ws, [22, 22, 22, 30])

    title_row(ws, 1, "장기보유특별공제 표 1·표 2 (v0.2 그대로 — v0.3-A 무변경)", span=4)

    section_header(ws, 3, "표 1 (일반과세) — 시행령 제159조의3", span=4)
    headers = ["보유연수", "공제율", "법령", "비고"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    table1 = [
        ("3년 이상 4년 미만", 0.06),
        ("4년 이상 5년 미만", 0.08),
        ("5년 이상 6년 미만", 0.10),
        ("6년 이상 7년 미만", 0.12),
        ("7년 이상 8년 미만", 0.14),
        ("8년 이상 9년 미만", 0.16),
        ("9년 이상 10년 미만", 0.18),
        ("10년 이상 11년 미만", 0.20),
        ("11년 이상 12년 미만", 0.22),
        ("12년 이상 13년 미만", 0.24),
        ("13년 이상 14년 미만", 0.26),
        ("14년 이상 15년 미만", 0.28),
        ("15년 이상", 0.30),
    ]
    for i, (yrs, rate) in enumerate(table1, start=5):
        calc_cell(ws, i, 1, yrs)
        calc_cell(ws, i, 2, rate, fmt="0%")
        calc_cell(ws, i, 3, "제159조의3", "")
        calc_cell(ws, i, 4, "")

    section_header(ws, 19, "표 2 (1세대1주택 + 12억 초과) — 시행령 제159조의4", span=4)
    headers2 = ["좌측 (보유)", "공제율", "우측 (거주)", "공제율"]
    for i, h in enumerate(headers2, start=1):
        c = ws.cell(row=20, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    table2 = [
        ("3년 이상 4년 미만", 0.12, "2년 이상 3년 미만 (보유 ≥ 3년)", 0.08),
        ("4년 이상 5년 미만", 0.16, "3년 이상 4년 미만", 0.12),
        ("5년 이상 6년 미만", 0.20, "4년 이상 5년 미만", 0.16),
        ("6년 이상 7년 미만", 0.24, "5년 이상 6년 미만", 0.20),
        ("7년 이상 8년 미만", 0.28, "6년 이상 7년 미만", 0.24),
        ("8년 이상 9년 미만", 0.32, "7년 이상 8년 미만", 0.28),
        ("9년 이상 10년 미만", 0.36, "8년 이상 9년 미만", 0.32),
        ("10년 이상", 0.40, "9년 이상 10년 미만", 0.36),
        ("", "", "10년 이상", 0.40),
    ]
    for i, (l1, r1, l2, r2) in enumerate(table2, start=21):
        calc_cell(ws, i, 1, l1)
        if r1 != "":
            calc_cell(ws, i, 2, r1, fmt="0%")
        calc_cell(ws, i, 3, l2)
        calc_cell(ws, i, 4, r2, fmt="0%")


# ============================================================
# 시트 7: 중과판정
# ============================================================
def build_sheet_heavy_judgment(wb):
    ws = wb.create_sheet("7_중과판정")
    setup_columns(ws, [25, 50, 25, 25])

    title_row(ws, 1, "다주택 중과 판정 4단계 조건 + issueFlag 카탈로그 (v0.3-A 신규)", span=4)

    section_header(ws, 3, "1. 4단계 조건 흐름도", span=4)
    headers = ["조건", "본문", "v0.X 활성", "법령"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    cond = [
        ("조건 1", "다주택 (householdHouseCount ≥ 2)", "v0.2 보존", "제104조 ⑦"),
        ("조건 2", "양도시 조정대상 (saleRegulated = true)", "v0.3-A 활성", "제104조 ⑦"),
        ("조건 3", "양도일 ≥ 2026-05-10", "v0.1 의사결정 #1", "중과 유예 종료"),
        ("조건 4", "비과세 미적용 (is1Se1House = false)", "v0.2 단계 2 결과", "제89조 ① ⅲ 미해당"),
    ]
    for i, (c1, c2, c3, c4) in enumerate(cond, start=5):
        label_cell(ws, i, 1, c1)
        calc_cell(ws, i, 2, c2)
        calc_cell(ws, i, 3, c3)
        calc_cell(ws, i, 4, c4)

    label_cell(ws, 9, 1, "AND 결과")
    heavy_cell(ws, 9, 2, "isHeavyTaxation = 조건1 ∧ 조건2 ∧ 조건3 ∧ 조건4")
    calc_cell(ws, 9, 3, "v0.3-A 신규")
    calc_cell(ws, 9, 4, "")

    section_header(ws, 11, "2. issueFlag 카탈로그 (v0.3-A 신규 7종)", span=4)
    headers2 = ["code", "severity", "발동 조건", "비고"]
    for i, h in enumerate(headers2, start=1):
        c = ws.cell(row=12, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    flags = [
        ("HEAVY_TAXATION_APPLIED", "warning", "isHeavyTaxation = true", "중과 발동 알림"),
        ("HEAVY_TAXATION_2_HOUSES", "info", "householdHouseCount = 2", "2주택 분기"),
        ("HEAVY_TAXATION_3_HOUSES", "info", "householdHouseCount ≥ 3", "3주택 이상 분기"),
        ("LONG_TERM_DEDUCTION_EXCLUDED_BY_MULTI_HOUSE_HEAVY", "info", "isHeavyTaxation = true", "장특공 배제"),
        ("SALE_REGULATED_USER_INPUT", "info", "saleRegulated 입력 시 항상", "B-033 책임 명시"),
        ("HEAVY_TAX_EXCLUSION_NOT_HANDLED", "info", "isHeavyTaxation = true 발동 시 항상", "시행령 제167조의10·11 단서 미처리"),
        ("HEAVY_TAX_TRANSITION_NOT_HANDLED", "info", "양도일 < 2026-05-10 + saleRegulated = true", "B-023 부칙 미반영"),
    ]
    for i, (code, sev, cond_text, note) in enumerate(flags, start=13):
        calc_cell(ws, i, 1, code)
        calc_cell(ws, i, 2, sev)
        calc_cell(ws, i, 3, cond_text)
        calc_cell(ws, i, 4, note)

    section_header(ws, 21, "3. 중과 시 단계 4·9 변경 요약", span=4)
    changes = [
        ("단계 4 (장특공)", "longTermDeduction = 0 (중과 시 표 1·2 모두 배제)", "제95조 ② 단서"),
        ("단계 9 (세율)", "calculatedTax = baseTax_with_addition + (taxBase − lowerBound) × (marginalRate + addition)", "동적 재계산"),
        ("§3-5 단서", "보유 < 2년 + 중과 시 max(단기세율 산출, 중과 누진 산출) 비교", "제104조 ⑦ 본문 단서"),
    ]
    for i, (stage, body, ref) in enumerate(changes, start=22):
        label_cell(ws, i, 1, stage)
        heavy_cell(ws, i, 2, body)
        calc_cell(ws, i, 3, ref)
        calc_cell(ws, i, 4, "")


# ============================================================
# 메인
# ============================================================
def main():
    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    build_sheet_rules(wb)
    build_sheet_tc011(wb)
    build_sheet_tc012(wb)
    build_sheet_tc013(wb)
    build_sheet_tc014(wb)
    build_sheet_rates(wb)
    build_sheet_long_term(wb)
    build_sheet_heavy_judgment(wb)

    # 출력
    out = "/home/claude/04_test_cases_manual_v03.xlsx"
    wb.save(out)
    print(f"✅ Saved: {out}")
    print(f"시트 수: {len(wb.sheetnames)}")
    print(f"시트 목록: {wb.sheetnames}")


if __name__ == "__main__":
    main()
